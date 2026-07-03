from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime, date
import logging
import os

from app.core.config import settings
from app.models.intelligence import Intelligence

logger = logging.getLogger(__name__)


class ExcelStorageService:
    """Excel存储服务"""
    
    def __init__(self):
        self.data_dir = Path(settings.excel_dir)
        self.data_dir.mkdir(parents=True, exist_ok=True)
        self._cache = {}  # 缓存: { "year-month": {"mtime": float, "data": List[Dict]} }
        self._stats_cache = {"mtime": 0, "data": None}
    
    def _get_excel_path(self, year: int, month: int) -> Path:
        """获取指定月份的Excel文件路径"""
        filename = f"intelligence_{year}-{month:02d}.xlsx"
        return self.data_dir / filename
    
    def _ensure_excel_exists(self, excel_path: Path) -> Workbook:
        """确保Excel文件存在，不存在则创建"""
        if excel_path.exists():
            return load_workbook(excel_path)
        
        # 创建新工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "情报数据"
        
        # 设置表头
        headers = ["日期", "行业", "标题", "摘要", "来源", "分类", "重要度", "关键词", "URL"]
        ws.append(headers)
        
        # 设置表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 设置列宽
        column_widths = [12, 15, 40, 50, 20, 10, 10, 30, 50]
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + idx)].width = width
        
        wb.save(excel_path)
        logger.info(f"创建新Excel文件: {excel_path}")
        return wb
    
    def save_intelligence(self, intelligence: Intelligence) -> bool:
        """保存单条情报到Excel"""
        try:
            intel_date = intelligence.date if isinstance(intelligence.date, date) else intelligence.date
            year, month = intel_date.year, intel_date.month
            
            excel_path = self._get_excel_path(year, month)
            wb = self._ensure_excel_exists(excel_path)
            ws = wb.active
            
            # 检查是否已存在（基于URL去重）
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[8] == intelligence.source_url:  # URL在第9列
                    logger.info(f"情报已存在，跳过: {intelligence.title[:30]}")
                    return False
            
            # 添加新行
            row_data = [
                intel_date.strftime("%Y-%m-%d"),
                intelligence.industry,
                intelligence.title,
                intelligence.summary,
                intelligence.source_name or "",
                intelligence.category,
                intelligence.importance,
                ", ".join(intelligence.keywords),
                intelligence.source_url
            ]
            ws.append(row_data)
            
            # 保存
            wb.save(excel_path)
            logger.info(f"保存情报成功: {intelligence.title[:30]}")
            return True
            
        except Exception as e:
            logger.error(f"保存情报失败: {str(e)}")
            return False
    
    def batch_save_intelligence(self, intelligence_list: List[Intelligence]) -> Dict[str, int]:
        """批量保存情报"""
        stats = {"success": 0, "skipped": 0, "failed": 0}
        
        for intel in intelligence_list:
            try:
                if self.save_intelligence(intel):
                    stats["success"] += 1
                else:
                    stats["skipped"] += 1
            except Exception as e:
                logger.error(f"批量保存失败: {str(e)}")
                stats["failed"] += 1
        
        logger.info(f"批量保存完成: {stats}")
        return stats
    
    def load_intelligence_by_month(
        self,
        year: int,
        month: int,
        industry: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """加载指定月份的情报"""
        excel_path = self._get_excel_path(year, month)
        
        if not excel_path.exists():
            return []
            
        try:
            mtime = excel_path.stat().st_mtime
            cache_key = f"{year}-{month}"
            
            # 使用缓存
            if cache_key in self._cache and self._cache[cache_key]["mtime"] == mtime:
                all_data = self._cache[cache_key]["data"]
            else:
                wb = load_workbook(excel_path, read_only=True)
                ws = wb.active
                
                all_data = []
                headers = [cell.value for cell in ws[1]]
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):  # 跳过空行
                        continue
                    all_data.append(dict(zip(headers, row)))
                    
                self._cache[cache_key] = {"mtime": mtime, "data": all_data}
                logger.info(f"更新缓存: {year}-{month:02d}, 共{len(all_data)}条")
            
            results = []
            for row_dict in all_data:
                # 筛选行业
                if industry and row_dict.get("行业") != industry:
                    continue
                results.append(row_dict)
            
            return results
            
        except Exception as e:
            logger.error(f"加载情报失败: {str(e)}")
            return []
    
    def load_intelligence_by_date_range(
        self,
        start_date: date,
        end_date: date,
        industry: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """加载日期范围内的情报"""
        all_results = []
        
        # 遍历日期范围内的所有月份
        current = start_date.replace(day=1)
        end = end_date.replace(day=1)
        
        while current <= end:
            month_results = self.load_intelligence_by_month(
                current.year,
                current.month,
                industry
            )
            
            # 过滤日期范围
            for result in month_results:
                date_str = result.get("日期")
                if date_str:
                    result_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                    if start_date <= result_date <= end_date:
                        all_results.append(result)
            
            # 移动到下一个月
            if current.month == 12:
                current = current.replace(year=current.year + 1, month=1)
            else:
                current = current.replace(month=current.month + 1)
        
        logger.info(f"加载日期范围情报: {start_date} 至 {end_date}, 共{len(all_results)}条")
        return all_results
    
    def get_statistics(self) -> Dict[str, Any]:
        """获取统计信息"""
        stats = {
            "total_files": 0,
            "total_intelligence": 0,
            "by_industry": {},
            "by_month": {}
        }
        
        try:
            # 计算所有文件的最新修改时间总和，作为缓存Key的依据
            files = list(self.data_dir.glob("intelligence_*.xlsx"))
            current_mtime = sum(f.stat().st_mtime for f in files)
            
            if self._stats_cache["mtime"] == current_mtime and self._stats_cache["data"] is not None:
                return self._stats_cache["data"]
            
            for excel_file in files:
                stats["total_files"] += 1
                
                wb = load_workbook(excel_file, read_only=True)
                ws = wb.active
                
                month_key = excel_file.stem.replace("intelligence_", "")
                month_count = 0
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    
                    month_count += 1
                    stats["total_intelligence"] += 1
                    
                    industry = row[1]  # 行业在第2列
                    if industry:
                        stats["by_industry"][industry] = stats["by_industry"].get(industry, 0) + 1
                
                stats["by_month"][month_key] = month_count
            
            self._stats_cache = {"mtime": current_mtime, "data": stats}
            logger.info(f"统计信息更新缓存: {stats}")
            return stats
            
        except Exception as e:
            logger.error(f"获取统计信息失败: {str(e)}")
            return stats


# 单例实例
_excel_service_instance = None

def get_excel_service() -> ExcelStorageService:
    """获取Excel服务单例"""
    global _excel_service_instance
    if _excel_service_instance is None:
        _excel_service_instance = ExcelStorageService()
    return _excel_service_instance